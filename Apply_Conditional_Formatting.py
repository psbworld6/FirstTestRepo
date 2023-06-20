from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Load the workbook
workbook = load_workbook('your_workbook.xlsx')

# Iterate through all sheets in the workbook
for sheet in workbook.sheetnames:
    # Select the current sheet
    worksheet = workbook[sheet]
    
    # Apply conditional formatting to columns F to O
    for col in range(6, 16):  # Columns F to O (6 to 15)
        # Retrieve unique values in the column
        unique_values = set(worksheet[get_column_letter(col)])
        
        # Apply conditional formatting to each unique value
        for value in unique_values:
            # Set fill and font color for unique values
            fill = PatternFill(start_color='FFFFD0', end_color='FFFFD0', fill_type='solid')
            font = Font(color='FF0000')
            
            # Iterate through cells in the column
            for cell in worksheet[get_column_letter(col)]:
                if cell.value == value:
                    # Apply formatting to the cell
                    cell.fill = fill
                    cell.font = font

# Save the modified workbook
workbook.save('your_workbook_formatted.xlsx')
