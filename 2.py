import os
from openpyxl import load_workbook

def apply_macro_to_files_in_directory(directory):
    # Specify the directory path where the Excel files are located
    folder_path = directory
    
    # Loop through each file in the directory
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        
        # Check if the file has .xls or .xlsx extension
        if filename.lower().endswith(('.xls', '.xlsx')):
            # Open the workbook
            workbook = load_workbook(filename=file_path)
            
            # Call the macro to apply the desired code
            your_macro_name(workbook)
            
            # Save and close the workbook
            workbook.save(filename=file_path)
            workbook.close()
            
            # Display the file name
            print(f"Macro applied to file: {filename}")
    
    print("Macro applied to all files in the directory.")

def your_macro_name(workbook):
    # Code for your macro goes here
    # You can use openpyxl methods to modify the workbook
    
    # Example: Display the names of all sheets in the workbook
    for sheet in workbook.sheetnames:
        print(sheet)

# Specify the directory path where the Excel files are located
directory_path = "C:/Your/Directory/Path"

# Call the function to apply the macro to files in the directory
apply_macro_to_files_in_directory(directory_path)
