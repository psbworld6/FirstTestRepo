import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font
import copy

# Directory path where the excel files are located
directory_path = "path/to/directory"

# Create a new workbook to store the copied sheets
consolidated_workbook = openpyxl.Workbook()

# List to store the sheet names
sheet_names = []

# Iterate over the files in the directory
for filename in os.listdir(directory_path):
    if filename.endswith(".xlsx"):
        # Extract the month and year from the file name
        month_year = filename.split("-")[0].strip()
        try:
            date_obj = datetime.strptime(month_year, '%B%Y')
            month_year_formatted = date_obj.strftime('%B%Y')

            # Open the source workbook and select the desired sheet
            source_workbook = openpyxl.load_workbook(os.path.join(directory_path, filename))
            source_sheet = source_workbook["Consolidated Test Report"]

            # Create a new sheet in the consolidated workbook
            new_sheet = consolidated_workbook.create_sheet(title=month_year_formatted)

            # Copy the sheet contents and formatting
            for row in source_sheet.iter_rows(values_only=True):
                new_sheet.append(row)

            for source_row, new_row in zip(source_sheet.iter_rows(), new_sheet.iter_rows()):
                for source_cell, new_cell in zip(source_row, new_row):
                    new_cell.font = copy.copy(source_cell.font)
                    new_cell.border = copy.copy(source_cell.border)
                    new_cell.fill = copy.copy(source_cell.fill)
                    new_cell.number_format = copy.copy(source_cell.number_format)
                    new_cell.alignment = copy.copy(source_cell.alignment)
                    new_cell.value = source_cell.value

            # Adjust row height to fit content
            for row in new_sheet.iter_rows():
                for cell in row:
                    new_sheet.row_dimensions[cell.row].height = source_sheet.row_dimensions[cell.row].height

            # Adjust column width to fit content
            for column in new_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = (max_length + 2) * 1.2
                new_sheet.column_dimensions[column_letter].width = adjusted_width

            # Align cells to the left
            for column in new_sheet.columns:
                for cell in column:
                    cell.alignment = cell.alignment.copy(horizontal="left")

            # Close the source workbook
            source_workbook.close()

            # Add the sheet name to the list
            sheet_names.append(month_year_formatted)
        except ValueError:
            continue

# Sort the sheet names in ascending order
sorted_sheet_names = sorted(sheet_names, key=lambda x: datetime.strptime(x, "%B%Y"))

# Reorder the sheets in the consolidated workbook based on the sorted sheet names
consolidated_workbook._sheets = [consolidated_workbook[sheet_name] for sheet_name in sorted_sheet_names]

# Save the consolidated workbook
consolidated_workbook.save(os.path.join(directory_path, "consolidated_reports.xlsx"))
consolidated_workbook.close()
