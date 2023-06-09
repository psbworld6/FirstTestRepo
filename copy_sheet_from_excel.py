import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter

# Get the directory path where the Excel files are located
directory = '/path/to/excel/files'  # Replace with the actual directory path

# Create a new workbook to store the consolidated sheets
consolidated_workbook = Workbook()

# Iterate over each file in the directory
for filename in os.listdir(directory):
    if filename.endswith(".xlsx"):
        # Load the Excel file
        file_path = os.path.join(directory, filename)
        workbook = load_workbook(file_path)

        # Check if the "Consolidated Test File" sheet exists
        if "Consolidated Test File" in workbook.sheetnames:
            # Get the reference to the source sheet
            source_sheet = workbook["Consolidated Test File"]

            # Create a new sheet in the consolidated workbook
            new_sheet = consolidated_workbook.create_sheet(title=filename)

            # Copy the sheet with formatting
            for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, min_col=1, max_col=source_sheet.max_column):
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.data_type = cell.data_type
                    new_cell.value = cell.value

                    # Copy the cell's style
                    if cell.has_style:
                        new_cell.font = Font(name=cell.font.name,
                                             size=cell.font.size,
                                             bold=cell.font.bold,
                                             italic=cell.font.italic,
                                             vertAlign=cell.font.vertAlign,
                                             underline=cell.font.underline,
                                             strike=cell.font.strike,
                                             color=cell.font.color)

                        new_cell.fill = PatternFill(fill_type=cell.fill.fill_type,
                                                    fgColor=cell.fill.fgColor,
                                                    bgColor=cell.fill.bgColor)

                        new_cell.border = Border(left=cell.border.left,
                                                 right=cell.border.right,
                                                 top=cell.border.top,
                                                 bottom=cell.border.bottom,
                                                 diagonal=cell.border.diagonal,
                                                 diagonal_direction=cell.border.diagonal_direction,
                                                 outline=cell.border.outline,
                                                 vertical=cell.border.vertical,
                                                 horizontal=cell.border.horizontal)

                        new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                       vertical=cell.alignment.vertical,
                                                       text_rotation=cell.alignment.text_rotation,
                                                       wrap_text=cell.alignment.wrap_text,
                                                       shrink_to_fit=cell.alignment.shrink_to_fit,
                                                       indent=cell.alignment.indent)

                        new_cell.number_format = cell.number_format

            # Rename the new sheet to the file name
            new_sheet.title = filename

print("Copying and renaming sheets completed.")

# Save the consolidated workbook to a new file
consolidated_file_path = os.path.join(directory, "consolidated_reports.xlsx")
consolidated_workbook.save(consolidated_file_path)

print(f"Consolidated sheets saved to {consolidated_file_path}.")
